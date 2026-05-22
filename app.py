from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Klient, Zaznam, Cinnost, Uloha
from datetime import datetime, timezone, timedelta
import bcrypt
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dochadzka-dev-key-zmenit')
database_url = os.environ.get('DATABASE_URL', 'sqlite:///dochadzka.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Pre prístup sa prosím prihláste.'
login_manager.login_message_category = 'error'

def now_sk():
    """Aktuálny čas v slovenskom časovom pásme (UTC+2 leto, UTC+1 zima)."""
    return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=2))).replace(tzinfo=None)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()


# ── AUTH ──────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/registracia', methods=['GET', 'POST'])
def registracia():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        meno   = request.form.get('meno', '').strip()
        email  = request.form.get('email', '').strip().lower()
        heslo  = request.form.get('heslo', '')
        heslo2 = request.form.get('heslo2', '')
        if not meno or not email or not heslo:
            flash('Vyplňte všetky polia.', 'error')
            return render_template('registracia.html')
        if heslo != heslo2:
            flash('Heslá sa nezhodujú.', 'error')
            return render_template('registracia.html')
        if len(heslo) < 6:
            flash('Heslo musí mať aspoň 6 znakov.', 'error')
            return render_template('registracia.html')
        if User.query.filter_by(email=email).first():
            flash('Email už existuje.', 'error')
            return render_template('registracia.html')
        hash_hesla = bcrypt.hashpw(heslo.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        is_admin = User.query.count() == 0
        user = User(meno=meno, email=email, heslo=hash_hesla, is_admin=is_admin)
        db.session.add(user)
        db.session.commit()
        flash('Registrácia úspešná! Prihláste sa.', 'success')
        return redirect(url_for('login'))
    return render_template('registracia.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        heslo = request.form.get('heslo', '')
        user  = User.query.filter_by(email=email).first()
        if user and bcrypt.checkpw(heslo.encode('utf-8'), user.heslo.encode('utf-8')):
            login_user(user, remember=True)
            return redirect(url_for('dashboard'))
        flash('Nesprávny email alebo heslo.', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    aktivny = Zaznam.query.filter_by(user_id=current_user.id, cas_stop=None)\
                          .order_by(Zaznam.cas_start.desc()).first()
    klienti = Klient.query.filter_by(user_id=current_user.id, aktivny=True)\
                          .order_by(Klient.nazov).all()
    posledne = Zaznam.query.filter_by(user_id=current_user.id)\
                           .filter(Zaznam.cas_stop.isnot(None))\
                           .order_by(Zaznam.cas_start.desc()).limit(5).all()
    # Úlohy pre mňa (neprebehnuté)
    moje_ulohy = Uloha.query.filter_by(prijimatel_id=current_user.id, hotovo=False)\
                             .order_by(Uloha.urgencia.desc(), Uloha.created_at.desc()).all()
    return render_template('dashboard.html', aktivny=aktivny, klienti=klienti,
                           posledne=posledne, moje_ulohy=moje_ulohy)


# ── ČASOVAČ ───────────────────────────────────────────────────────────────────

@app.route('/start', methods=['POST'])
@login_required
def start_cas():
    if Zaznam.query.filter_by(user_id=current_user.id, cas_stop=None).first():
        flash('Máte aktívny časovač. Najprv ho zastavte.', 'error')
        return redirect(url_for('dashboard'))
    klient_id = request.form.get('klient_id')
    klient = Klient.query.filter_by(id=klient_id, user_id=current_user.id).first()
    if not klient:
        flash('Vyberte platného klienta.', 'error')
        return redirect(url_for('dashboard'))
    now = now_sk()
    zaznam = Zaznam(user_id=current_user.id, klient_id=klient.id, datum=now.date(), cas_start=now)
    db.session.add(zaznam)
    db.session.commit()
    flash(f'Časovač spustený — {klient.nazov}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/stop', methods=['POST'])
@login_required
def stop_cas():
    zaznam_id = request.form.get('zaznam_id')
    zaznam = Zaznam.query.filter_by(id=zaznam_id, user_id=current_user.id, cas_stop=None).first()
    if not zaznam:
        flash('Záznam nenájdený.', 'error')
        return redirect(url_for('dashboard'))
    now = now_sk()
    zaznam.cas_stop = now
    posledna = Cinnost.query.filter_by(zaznam_id=zaznam.id, cas_do=None).first()
    if posledna:
        posledna.cas_do = now
    poznamka = request.form.get('poznamka', '').strip()
    if poznamka:
        zaznam.poznamka = poznamka
    db.session.commit()
    flash(f'Práca ukončená — {zaznam.trvanie_format}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/cinnost/pridat', methods=['POST'])
@login_required
def pridat_cinnost():
    zaznam_id  = request.form.get('zaznam_id')
    popis      = request.form.get('popis', '').strip()
    redirect_to = request.form.get('redirect', 'dashboard')
    if not popis:
        flash('Zadajte popis činnosti.', 'error')
        return redirect(url_for(redirect_to))
    zaznam = Zaznam.query.filter_by(id=zaznam_id, user_id=current_user.id).first()
    if not zaznam:
        flash('Záznam nenájdený.', 'error')
        return redirect(url_for(redirect_to))
    now = now_sk()
    posledna = Cinnost.query.filter_by(zaznam_id=zaznam.id, cas_do=None).first()
    if posledna:
        posledna.cas_do = now
    # Ak je záznam už ukončený, čas činnosti = teraz (len zápis)
    cinnost = Cinnost(
        zaznam_id=zaznam.id,
        popis=popis,
        cas_od=now,
        cas_do=now if zaznam.cas_stop is not None else None
    )
    db.session.add(cinnost)
    db.session.commit()
    flash('Činnosť pridaná.', 'success')
    return redirect(url_for(redirect_to))

@app.route('/zaznam/<int:zaznam_id>/poznamka', methods=['POST'])
@login_required
def ulozit_poznamku(zaznam_id):
    zaznam = Zaznam.query.filter_by(id=zaznam_id, user_id=current_user.id).first_or_404()
    zaznam.poznamka = request.form.get('poznamka', '').strip()
    db.session.commit()
    flash('Poznámka uložená.', 'success')
    return redirect(url_for('vykaz'))


# ── KLIENTI ───────────────────────────────────────────────────────────────────

@app.route('/klienti')
@login_required
def klienti():
    zoznam = Klient.query.filter_by(user_id=current_user.id).order_by(Klient.nazov).all()
    return render_template('klienti.html', klienti=zoznam)

@app.route('/klient/pridat', methods=['POST'])
@login_required
def pridat_klienta():
    nazov = request.form.get('nazov', '').strip()
    if not nazov:
        flash('Zadajte názov klienta.', 'error')
        return redirect(url_for('klienti'))
    if Klient.query.filter_by(user_id=current_user.id, nazov=nazov).first():
        flash('Klient s týmto názvom už existuje.', 'error')
        return redirect(url_for('klienti'))
    db.session.add(Klient(nazov=nazov, user_id=current_user.id))
    db.session.commit()
    flash(f'Klient „{nazov}" pridaný.', 'success')
    return redirect(url_for('klienti'))

@app.route('/klient/<int:klient_id>/upravit', methods=['POST'])
@login_required
def upravit_klienta(klient_id):
    klient = Klient.query.filter_by(id=klient_id, user_id=current_user.id).first_or_404()
    novy_nazov = request.form.get('nazov', '').strip()
    if not novy_nazov:
        flash('Názov nemôže byť prázdny.', 'error')
        return redirect(url_for('klienti'))
    klient.nazov = novy_nazov
    db.session.commit()
    flash(f'Klient premenovaný na „{novy_nazov}".', 'success')
    return redirect(url_for('klienti'))

@app.route('/klient/<int:klient_id>/archivovat', methods=['POST'])
@login_required
def archivovat_klienta(klient_id):
    klient = Klient.query.filter_by(id=klient_id, user_id=current_user.id).first_or_404()
    klient.aktivny = not klient.aktivny
    db.session.commit()
    flash('Klient ' + ('aktivovaný.' if klient.aktivny else 'archivovaný.'), 'success')
    return redirect(url_for('klienti'))


# ── VÝKAZ ─────────────────────────────────────────────────────────────────────

def _filter_zaznamy(user_id, datum_od, datum_do, klient_id):
    q = Zaznam.query.filter_by(user_id=user_id).filter(Zaznam.cas_stop.isnot(None))
    if datum_od:
        try:
            q = q.filter(Zaznam.datum >= datetime.strptime(datum_od, '%Y-%m-%d').date())
        except: pass
    if datum_do:
        try:
            q = q.filter(Zaznam.datum <= datetime.strptime(datum_do, '%Y-%m-%d').date())
        except: pass
    if klient_id:
        q = q.filter(Zaznam.klient_id == klient_id)
    return q

@app.route('/vykaz')
@login_required
def vykaz():
    datum_od  = request.args.get('datum_od', '')
    datum_do  = request.args.get('datum_do', '')
    klient_id = request.args.get('klient_id', '')
    zaznamy   = _filter_zaznamy(current_user.id, datum_od, datum_do, klient_id)\
                    .order_by(Zaznam.datum.desc(), Zaznam.cas_start.desc()).all()
    klienti   = Klient.query.filter_by(user_id=current_user.id).order_by(Klient.nazov).all()
    celk_sek  = sum(z.trvanie_sekundy or 0 for z in zaznamy)
    celkovy   = f'{celk_sek // 3600:02d}:{(celk_sek % 3600) // 60:02d}'
    return render_template('vykaz.html', zaznamy=zaznamy, klienti=klienti,
                           celkovy=celkovy, datum_od=datum_od, datum_do=datum_do,
                           klient_id=klient_id)

@app.route('/vykaz/export')
@login_required
def export_excel():
    datum_od  = request.args.get('datum_od', '')
    datum_do  = request.args.get('datum_do', '')
    klient_id = request.args.get('klient_id', '')
    zaznamy   = _filter_zaznamy(current_user.id, datum_od, datum_do, klient_id)\
                    .order_by(Zaznam.datum.asc(), Zaznam.cas_start.asc()).all()
    output = _build_excel(zaznamy, current_user.meno)
    filename = f'dochadzka_{current_user.meno.replace(" ","_")}_{now_sk().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── ADMIN ─────────────────────────────────────────────────────────────────────

@app.route('/admin')
@login_required
def admin():
    if not current_user.is_admin:
        flash('Prístup zamietnutý.', 'error')
        return redirect(url_for('dashboard'))
    users = User.query.order_by(User.meno).all()
    return render_template('admin.html', users=users)

@app.route('/admin/vykaz/<int:user_id>')
@login_required
def admin_vykaz(user_id):
    if not current_user.is_admin:
        flash('Prístup zamietnutý.', 'error')
        return redirect(url_for('dashboard'))
    user      = User.query.get_or_404(user_id)
    datum_od  = request.args.get('datum_od', '')
    datum_do  = request.args.get('datum_do', '')
    klient_id = request.args.get('klient_id', '')
    zaznamy   = _filter_zaznamy(user_id, datum_od, datum_do, klient_id)\
                    .order_by(Zaznam.datum.desc(), Zaznam.cas_start.desc()).all()
    klienti   = Klient.query.filter_by(user_id=user_id).order_by(Klient.nazov).all()
    celk_sek  = sum(z.trvanie_sekundy or 0 for z in zaznamy)
    celkovy   = f'{celk_sek // 3600:02d}:{(celk_sek % 3600) // 60:02d}'
    return render_template('admin_vykaz.html', zamestnanec=user, zaznamy=zaznamy,
                           klienti=klienti, celkovy=celkovy,
                           datum_od=datum_od, datum_do=datum_do, klient_id=klient_id)

@app.route('/admin/export/<int:user_id>')
@login_required
def admin_export(user_id):
    if not current_user.is_admin:
        flash('Prístup zamietnutý.', 'error')
        return redirect(url_for('dashboard'))
    user      = User.query.get_or_404(user_id)
    datum_od  = request.args.get('datum_od', '')
    datum_do  = request.args.get('datum_do', '')
    klient_id = request.args.get('klient_id', '')
    zaznamy   = _filter_zaznamy(user_id, datum_od, datum_do, klient_id)\
                    .order_by(Zaznam.datum.asc(), Zaznam.cas_start.asc()).all()
    output   = _build_excel(zaznamy, user.meno)
    filename = f'dochadzka_{user.meno.replace(" ","_")}_{now_sk().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/toggle_admin/<int:user_id>', methods=['POST'])
@login_required
def toggle_admin(user_id):
    if not current_user.is_admin:
        flash('Prístup zamietnutý.', 'error')
        return redirect(url_for('dashboard'))
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Nemôžete zmeniť vlastnú rolu.', 'error')
        return redirect(url_for('admin'))
    user.is_admin = not user.is_admin
    db.session.commit()
    flash(f'{user.meno} — rola zmenená.', 'success')
    return redirect(url_for('admin'))


# ── ÚLOHY ─────────────────────────────────────────────────────────────────────

@app.route('/ulohy')
@login_required
def ulohy():
    # Úlohy ktoré som ja poslal
    odoslane = Uloha.query.filter_by(odosielatel_id=current_user.id)\
                          .order_by(Uloha.hotovo.asc(), Uloha.urgencia.desc(), Uloha.created_at.desc()).all()
    # Úlohy pre mňa
    prijate = Uloha.query.filter_by(prijimatel_id=current_user.id)\
                         .order_by(Uloha.hotovo.asc(), Uloha.urgencia.desc(), Uloha.created_at.desc()).all()
    users = User.query.filter(User.id != current_user.id).order_by(User.meno).all()
    return render_template('ulohy.html', odoslane=odoslane, prijate=prijate, users=users)

@app.route('/uloha/pridat', methods=['POST'])
@login_required
def pridat_ulohu():
    prijimatel_id = request.form.get('prijimatel_id')
    text          = request.form.get('text', '').strip()
    urgencia      = int(request.form.get('urgencia', 2))
    if not text or not prijimatel_id:
        flash('Vyplňte všetky polia.', 'error')
        return redirect(url_for('ulohy'))
    prijimatel = User.query.get(prijimatel_id)
    if not prijimatel:
        flash('Príjemca nenájdený.', 'error')
        return redirect(url_for('ulohy'))
    uloha = Uloha(
        odosielatel_id=current_user.id,
        prijimatel_id=int(prijimatel_id),
        text=text,
        urgencia=urgencia,
        created_at=now_sk()
    )
    db.session.add(uloha)
    db.session.commit()
    flash(f'Úloha odoslaná pre {prijimatel.meno}.', 'success')
    return redirect(url_for('ulohy'))

@app.route('/uloha/<int:uloha_id>/hotovo', methods=['POST'])
@login_required
def oznacit_hotovo(uloha_id):
    uloha = Uloha.query.filter_by(id=uloha_id, prijimatel_id=current_user.id).first_or_404()
    uloha.hotovo = not uloha.hotovo
    uloha.hotovo_at = now_sk() if uloha.hotovo else None
    db.session.commit()
    return redirect(url_for('ulohy'))


# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────

def _build_excel(zaznamy, meno):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dochádzka'

    navy  = '1E3A5F'
    blue2 = '2E6DA4'
    white = 'FFFFFF'
    light = 'EEF4FB'

    hdr_font = Font(bold=True, color=white, name='Calibri', size=11)
    hdr_fill = PatternFill('solid', start_color=navy)
    alt_fill = PatternFill('solid', start_color=light)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin = lambda: Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'))

    ws.merge_cells('A1:I1')
    ws['A1'] = f'Výkaz dochádzky — {meno}'
    ws['A1'].font = Font(bold=True, size=15, color=navy, name='Calibri')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Exportované: {now_sk().strftime("%d.%m.%Y %H:%M")}'
    ws['A2'].font = Font(size=9, color='888888', name='Calibri')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 15

    cols   = ['Dátum', 'Klient', 'Začiatok', 'Koniec', 'Celkový čas', 'Činnosť', 'Od', 'Do', 'Poznámka']
    widths = [14, 22, 11, 11, 13, 38, 11, 11, 28]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    row = 4
    for i, z in enumerate(zaznamy):
        cinnosti = z.cinnosti if z.cinnosti else [None]
        fill     = alt_fill if (i % 2 == 0) else PatternFill('solid', start_color=white)
        for ji, cin in enumerate(cinnosti):
            for ci in range(1, 10):
                c = ws.cell(row=row, column=ci)
                c.border = thin(); c.fill = fill; c.alignment = left
                c.font = Font(name='Calibri', size=10)
            if ji == 0:
                ws.cell(row=row, column=1).value = z.datum.strftime('%d.%m.%Y') if z.datum else ''
                ws.cell(row=row, column=1).alignment = center
                ws.cell(row=row, column=2).value = z.klient.nazov if z.klient else ''
                ws.cell(row=row, column=3).value = z.cas_start.strftime('%H:%M') if z.cas_start else ''
                ws.cell(row=row, column=3).alignment = center
                ws.cell(row=row, column=4).value = z.cas_stop.strftime('%H:%M') if z.cas_stop else ''
                ws.cell(row=row, column=4).alignment = center
                ws.cell(row=row, column=5).value = z.trvanie_format
                ws.cell(row=row, column=5).alignment = center
                ws.cell(row=row, column=5).font = Font(name='Calibri', size=10, bold=True)
                ws.cell(row=row, column=9).value = z.poznamka or ''
            if cin:
                ws.cell(row=row, column=6).value = cin.popis
                ws.cell(row=row, column=7).value = cin.cas_od.strftime('%H:%M') if cin.cas_od else ''
                ws.cell(row=row, column=7).alignment = center
                ws.cell(row=row, column=8).value = cin.cas_do.strftime('%H:%M') if cin.cas_do else ''
                ws.cell(row=row, column=8).alignment = center
            ws.row_dimensions[row].height = 18
            row += 1

    row += 1
    celk_sek = sum(z.trvanie_sekundy or 0 for z in zaznamy)
    celk_str = f'{celk_sek // 3600:02d}:{(celk_sek % 3600) // 60:02d}'
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1).value = 'CELKOVÝ ČAS:'
    ws.cell(row=row, column=1).font  = Font(bold=True, name='Calibri', size=11, color=white)
    ws.cell(row=row, column=1).fill  = PatternFill('solid', start_color=navy)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=row, column=5).value = celk_str
    ws.cell(row=row, column=5).font  = Font(bold=True, name='Calibri', size=12, color=white)
    ws.cell(row=row, column=5).fill  = PatternFill('solid', start_color=blue2)
    ws.cell(row=row, column=5).alignment = center
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
